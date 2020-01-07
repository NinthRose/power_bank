import csv
import os

import openpyxl

from orca import data_dir, export_task_dir


def write_file(username, data_set, filename, bytes):
    file_path = os.path.join(data_dir, username, data_set, filename)
    if not os.path.isdir(os.path.dirname(file_path)):
        os.makedirs(os.path.dirname(file_path))
    with open(file_path, 'wb') as f:
        f.write(bytes)
    return file_path


def __write_sheet(sheet, title, content):
    assert isinstance(title, str), 'title should be string or chars.'
    sheet.title = title
    for i, line in enumerate(content):
        for j, cell in enumerate(line):
            sheet.cell(i + 1, j + 1, cell)


def __write_csv(csv_writer, title, content):
    csv_writer.writerow([title])
    for line in content:
        csv_writer.writerow(line)
    csv_writer.writerow([])


def export_table(username, filename, data, is_csv=None):
    assert data and isinstance(data, dict), \
        'illegal params: the key of data should be the name of sheet:{}'.format(type(data))
    for i in data.values():
        assert isinstance(i, dict), 'illegal params: the value of data should be the dict of span(label:[span])'
    file_path = os.path.join(export_task_dir, username, filename)
    suffix = '.csv' if is_csv else '.xlsx'
    if not file_path.endswith(suffix):
        file_path += suffix

    if is_csv:
        output_stream = open(file_path, mode="w", encoding="utf-8", newline='')
        csv_writer = csv.writer(output_stream, quotechar='"', quoting=csv.QUOTE_NONE, escapechar='"')
    else:
        wb = openpyxl.Workbook()
        sheet = wb.active
    content = list()
    content.append(['doc_amount', len(data)])
    spans = list(data.values())[0]
    labels = list(spans.keys())
    content.append(['label_amount', len(labels)])
    content.append(['label_list', *labels])
    if is_csv:
        __write_csv(csv_writer, 'info', content)
    else:
        __write_sheet(sheet, 'info', content)
    for name, spans in data.items():
        content = list()
        content.append(['label', 'name', 'start', 'end'])
        for label, span_list in spans.items():
            for s in span_list:
                content.append([label, s['name'], s['start'], s['end']])

        if is_csv:
            __write_csv(csv_writer, str(name), content)
        else:
            sheet = wb.create_sheet()
            __write_sheet(sheet, str(name), content)

    dir, _ = os.path.split(file_path)
    if not os.path.isdir(dir):
        os.makedirs(dir)

    if is_csv:
        output_stream.close()
    else:
        wb.save(file_path)
    return file_path
